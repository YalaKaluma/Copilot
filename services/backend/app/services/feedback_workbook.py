"""Generate a reviewer-friendly Excel export for a saved conversation."""

from __future__ import annotations

import json
from io import BytesIO
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

if TYPE_CHECKING:
    from packages.db.models.conversation import Conversation


def _text(value: Any) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = json.dumps(value, ensure_ascii=False, default=str)
    value = value[:32_000]
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def build_feedback_workbook(conversation: Conversation) -> bytes:
    workbook = Workbook()
    review = workbook.active
    review.title = "Response Review"
    log = workbook.create_sheet("Conversation Log")

    headers = [
        "Review ID", "Session ID", "Timestamp", "User question",
        "Assistant response", "Conversation context", "Tool / metadata",
        "Plan", "Rating (1-5)", "Issue category", "Reviewer comments",
        "Suggested better answer", "Prompt / guideline improvement",
        "Review status",
    ]
    review.append(headers)
    transcript: list[str] = []
    question = ""
    response_number = 0
    for message in conversation.messages:
        transcript.append(f"{message.role.upper()}: {message.content}")
        if message.role == "user":
            question = message.content
            continue
        if message.role != "assistant":
            continue
        response_number += 1
        metadata = message.message_metadata or {}
        review.append([
            f"R{response_number:04d}", conversation.session_id,
            message.created_at.isoformat(), question, message.content,
            "\n\n".join(transcript), _text(metadata),
            _text(conversation.plan_data), "", "", "", "", "",
            "Not reviewed",
        ])

    dark = PatternFill("solid", fgColor="17151A")
    editable = PatternFill("solid", fgColor="FFF4D6")
    for cell in review[1]:
        cell.fill = dark
        cell.font = Font(bold=True, color="FFFFFF")
    for row in review.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row[8:14]:
            cell.fill = editable
    widths = [12, 24, 22, 42, 70, 75, 45, 55, 14, 25, 50, 55, 55, 18]
    for index, width in enumerate(widths, 1):
        review.column_dimensions[review.cell(1, index).column_letter].width = width
    review.freeze_panes = "D2"
    review.auto_filter.ref = f"A1:N{max(review.max_row, 2)}"

    rating = DataValidation(type="list", formula1='"1,2,3,4,5"')
    category = DataValidation(
        type="list",
        formula1='"None,Incorrect answer,Missing evidence,Wrong endpoint,Bad scope,Too verbose,Too brief,Other"',
    )
    status = DataValidation(
        type="list", formula1='"Not reviewed,Reviewed,Action required"'
    )
    for validation, cells in (
        (rating, "I2:I1000"), (category, "J2:J1000"), (status, "N2:N1000")
    ):
        review.add_data_validation(validation)
        validation.add(cells)

    log.append(["Message #", "Timestamp", "Role", "Message", "Metadata"])
    for index, message in enumerate(conversation.messages, 1):
        log.append([
            index, message.created_at.isoformat(), message.role,
            message.content, _text(message.message_metadata),
        ])
    for cell in log[1]:
        cell.fill = dark
        cell.font = Font(bold=True, color="FFFFFF")
    for row in log.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column, width in zip("ABCDE", [12, 22, 14, 90, 55]):
        log.column_dimensions[column].width = width
    log.freeze_panes = "D2"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
