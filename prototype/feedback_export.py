"""Create an Excel workbook for reviewing copilot conversations."""

from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


INK = "17151A"
WHITE = "FFFFFF"
MUTED = "756E74"
BURGUNDY = "81264A"
BURGUNDY_SOFT = "F4E9EE"
TEAL = "168B7D"
LIGHT = "F7F5F4"
LINE = "DDD7DA"


def _safe_text(value: Any, limit: int = 32_000) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = json.dumps(value, ensure_ascii=False, default=str)
    value = value[:limit]
    if value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _timestamp(message: dict[str, Any]) -> str:
    return _safe_text(message.get("timestamp"))


def build_feedback_workbook(
    messages: list[dict[str, Any]],
    *,
    session_id: str,
    tenant_code: str | None,
    model: str,
) -> bytes:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    review = wb.create_sheet("Response Review")
    log = wb.create_sheet("Conversation Log")

    for sheet in (instructions, review, log):
        sheet.sheet_view.showGridLines = False

    instructions.merge_cells("A1:F1")
    instructions["A1"] = "SKAI Growth Copilot — Feedback Workbook"
    instructions["A1"].font = Font(size=18, bold=True, color=WHITE)
    instructions["A1"].fill = PatternFill("solid", fgColor=INK)
    instructions["A1"].alignment = Alignment(vertical="center")
    instructions.row_dimensions[1].height = 34
    instruction_rows = [
        ("Purpose", "Review each assistant response and provide structured feedback for prompt and guidance improvement."),
        ("How to review", "Use the Response Review sheet. Enter a 1–5 rating, select an issue category, and add comments where useful."),
        ("Rating scale", "5 = excellent; 4 = good/minor edits; 3 = partly useful; 2 = substantial problems; 1 = incorrect or unusable."),
        ("Comments", "Describe what was missing, misleading, too verbose, or poorly scoped. Be specific about the desired behavior."),
        ("Suggested answer", "Optionally write or outline the answer you would have preferred."),
        ("Improvement proposal", "Optionally identify a prompt, endpoint-routing, tool-description, or analytical-guideline change."),
        ("Return loop", "Save the completed workbook and provide it back for analysis. Ratings and comments can then be grouped into recurring failure patterns."),
        ("Session ID", session_id),
        ("Tenant", tenant_code or "Default tenant"),
        ("Model", model),
    ]
    for row_index, (label, detail) in enumerate(instruction_rows, start=3):
        instructions.cell(row_index, 1, label)
        instructions.cell(row_index, 2, detail)
        instructions.cell(row_index, 1).font = Font(bold=True, color=BURGUNDY)
        instructions.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 100

    review_headers = [
        "Review ID",
        "Session ID",
        "Timestamp",
        "Tenant",
        "Model",
        "User question",
        "Assistant response",
        "Conversation context",
        "Tool",
        "Tool arguments",
        "Plan / limitation",
        "Rating (1-5)",
        "Issue category",
        "Reviewer comments",
        "Suggested better answer",
        "Prompt / guideline improvement",
        "Review status",
    ]
    review.append(review_headers)
    transcript: list[str] = []
    review_number = 0
    last_user_question = ""
    for message in messages:
        role = message.get("role", "unknown")
        content = _safe_text(message.get("content"))
        transcript.append(f"{role.upper()}: {content}")
        if role == "user":
            last_user_question = content
            continue
        if role != "assistant":
            continue
        review_number += 1
        plan = message.get("plan") or {}
        review.append(
            [
                f"R{review_number:04d}",
                session_id,
                _timestamp(message),
                tenant_code or "default",
                model,
                last_user_question,
                content,
                _safe_text("\n\n".join(transcript)),
                _safe_text(plan.get("tool")),
                _safe_text(plan.get("arguments")),
                _safe_text(
                    {
                        "interpretation": plan.get("interpretation"),
                        "steps": plan.get("steps"),
                        "limitation": plan.get("limitation"),
                    }
                ),
                "",
                "",
                "",
                "",
                "",
                "Not reviewed",
            ]
        )

    review.freeze_panes = "F2"
    review.auto_filter.ref = f"A1:Q{max(review.max_row, 2)}"
    review.row_dimensions[1].height = 32
    header_fill = PatternFill("solid", fgColor=INK)
    for cell in review[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    editable_fill = PatternFill("solid", fgColor="FFF4D6")
    for row in review.iter_rows(min_row=2, min_col=12, max_col=17):
        for cell in row:
            cell.fill = editable_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in review.iter_rows(min_row=2, max_row=review.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        review.row_dimensions[row[0].row].height = 78

    widths = [12, 22, 20, 18, 16, 42, 65, 70, 24, 45, 50, 14, 24, 48, 55, 55, 18]
    for index, width in enumerate(widths, start=1):
        review.column_dimensions[chr(64 + index)].width = width

    rating_validation = DataValidation(
        type="list", formula1='"1,2,3,4,5"', allow_blank=True
    )
    issue_validation = DataValidation(
        type="list",
        formula1='"None,Incorrect answer,Missing evidence,Wrong tool or endpoint,Bad scope or filters,Unsupported claim,Too verbose,Too brief,Unclear structure,Other"',
        allow_blank=True,
    )
    status_validation = DataValidation(
        type="list", formula1='"Not reviewed,Reviewed,Action required"'
    )
    review.add_data_validation(rating_validation)
    review.add_data_validation(issue_validation)
    review.add_data_validation(status_validation)
    rating_validation.add(f"L2:L{max(review.max_row, 1000)}")
    issue_validation.add(f"M2:M{max(review.max_row, 1000)}")
    status_validation.add(f"Q2:Q{max(review.max_row, 1000)}")
    review.conditional_formatting.add(
        f"L2:L{max(review.max_row, 1000)}",
        CellIsRule(operator="lessThanOrEqual", formula=["2"], fill=PatternFill("solid", fgColor="F4CCCC")),
    )
    review.conditional_formatting.add(
        f"L2:L{max(review.max_row, 1000)}",
        CellIsRule(operator="greaterThanOrEqual", formula=["4"], fill=PatternFill("solid", fgColor="D9EAD3")),
    )

    log_headers = ["Session ID", "Message #", "Timestamp", "Tenant", "Role", "Message", "Tool", "Plan"]
    log.append(log_headers)
    for index, message in enumerate(messages, start=1):
        plan = message.get("plan") or {}
        log.append(
            [
                session_id,
                index,
                _timestamp(message),
                tenant_code or "default",
                _safe_text(message.get("role")),
                _safe_text(message.get("content")),
                _safe_text(plan.get("tool")),
                _safe_text(plan),
            ]
        )
    log.freeze_panes = "F2"
    log.auto_filter.ref = f"A1:H{max(log.max_row, 2)}"
    for cell in log[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color=WHITE)
    for row in log.iter_rows(min_row=2, max_row=log.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        log.row_dimensions[row[0].row].height = 58
    for column, width in zip("ABCDEFGH", [22, 12, 20, 18, 14, 80, 25, 65]):
        log.column_dimensions[column].width = width

    thin = Side(style="thin", color=LINE)
    for sheet in (review, log):
        used = sheet.iter_rows(min_row=1, max_row=sheet.max_row)
        for row in used:
            for cell in row:
                cell.border = Border(bottom=thin)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def build_hypothesis_feedback_workbook(
    hypotheses: list[dict[str, Any]],
    *,
    scope: dict[str, Any],
    source_errors: dict[str, str],
    raw_evidence: dict[str, Any],
    tenant_code: str | None,
    model: str,
) -> bytes:
    """Create a two-level review workbook for hypotheses and evidence cards."""
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    hypothesis_review = wb.create_sheet("Hypothesis Review")
    evidence_review = wb.create_sheet("Evidence Review")
    source_status = wb.create_sheet("Source Status")

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    instructions.merge_cells("A1:F1")
    instructions["A1"] = "SKAI Pricing Hypothesis - Feedback Workbook"
    instructions["A1"].font = Font(size=18, bold=True, color=WHITE)
    instructions["A1"].fill = PatternFill("solid", fgColor=INK)
    instructions["A1"].alignment = Alignment(vertical="center")
    instructions.row_dimensions[1].height = 34
    guidance = [
        ("Purpose", "Review the agent's pricing hypotheses and every supporting or counterevidence card."),
        ("Hypothesis Review", "Rate whether each overall hypothesis is commercially sound, correctly scoped, and actionable."),
        ("Evidence Review", "Review each evidence card independently. Check factual validity, relevance, direction, interpretation, and source use."),
        ("Rating scale", "5 = correct and useful; 4 = good/minor edit; 3 = partly useful; 2 = materially flawed; 1 = incorrect or unusable."),
        ("Suggested correction", "Where possible, describe the hypothesis or evidence interpretation the agent should have produced."),
        ("Feedback loop", "Return the completed workbook so feedback can be translated into prompt, tool-routing, and analytical-guideline improvements."),
        ("Tenant", tenant_code or "Default tenant"),
        ("Model", model),
        ("Scope", _safe_text(scope)),
    ]
    for row_index, (label, detail) in enumerate(guidance, start=3):
        instructions.cell(row_index, 1, label)
        instructions.cell(row_index, 2, detail)
        instructions.cell(row_index, 1).font = Font(bold=True, color=BURGUNDY)
        instructions.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 105

    hypothesis_headers = [
        "Hypothesis ID", "Lever", "Direction", "Statement", "Proposed opportunity",
        "Evidence status", "Confidence", "Priority", "Estimated value", "Value basis",
        "Brand", "SKU", "Retailer", "Overall rating (1-5)", "Reviewer decision",
        "Hypothesis feedback", "Suggested corrected hypothesis", "Missing analysis",
        "Review status",
    ]
    hypothesis_review.append(hypothesis_headers)
    for index, hypothesis in enumerate(hypotheses, start=1):
        hypothesis_review.append([
            hypothesis.get("id") or f"H-{index}", scope.get("lever", "Pricing"),
            hypothesis.get("direction"), hypothesis.get("statement"),
            hypothesis.get("opportunity"), hypothesis.get("evidence_status"),
            hypothesis.get("confidence"), hypothesis.get("priority"),
            hypothesis.get("estimated_value"), hypothesis.get("value_basis"),
            scope.get("brand"), scope.get("sku"), scope.get("retailer"),
            "", "", "", "", "", "Not reviewed",
        ])

    evidence_headers = [
        "Evidence ID", "Hypothesis ID", "Direction tested", "Evidence role",
        "Finding", "Agent interpretation", "Strength", "Analytical source",
        "Evidence scope", "Brand", "SKU", "Retailer", "Validity rating (1-5)",
        "Relevance rating (1-5)", "Direction correct?", "Source used correctly?",
        "Evidence feedback", "Corrected finding / interpretation", "Keep or remove",
        "Review status",
    ]
    evidence_review.append(evidence_headers)
    evidence_number = 0
    for hypothesis_index, hypothesis in enumerate(hypotheses, start=1):
        hypothesis_id = hypothesis.get("id") or f"H-{hypothesis_index}"
        for evidence in hypothesis.get("evidence", []):
            evidence_number += 1
            evidence_review.append([
                f"E-{evidence_number:04d}", hypothesis_id,
                hypothesis.get("direction"), evidence.get("direction"),
                evidence.get("finding"), evidence.get("interpretation"),
                evidence.get("strength"), evidence.get("source"), evidence.get("scope"),
                scope.get("brand"), scope.get("sku"), scope.get("retailer"),
                "", "", "", "", "", "", "", "Not reviewed",
            ])

    source_status.append(["Source", "Status", "Error / limitation", "Raw response"])
    expected_sources = {
        "market_landscape_overall", "market_landscape_selected_retailer",
        "price_ladder_overall", "price_ladder_selected_retailer", "price_pack_curve",
    }
    all_sources = sorted(
        expected_sources | set(source_errors) | set(raw_evidence)
        - {"source_errors"}
    )
    for source in all_sources:
        status = "Unavailable" if source in source_errors else "Available" if source in raw_evidence else "Not requested"
        source_status.append([
            source, status, source_errors.get(source, ""),
            _safe_text(raw_evidence.get(source)),
        ])

    header_fill = PatternFill("solid", fgColor=INK)
    editable_fill = PatternFill("solid", fgColor="FFF4D6")
    thin = Side(style="thin", color=LINE)
    for sheet, editable_start, widths in (
        (hypothesis_review, 14, [16, 14, 18, 48, 55, 18, 12, 10, 18, 48, 20, 30, 22, 16, 20, 55, 55, 45, 18]),
        (evidence_review, 14, [14, 16, 18, 20, 55, 62, 14, 24, 42, 20, 30, 22, 16, 18, 18, 22, 55, 60, 18, 18]),
        (source_status, 99, [38, 16, 65, 100]),
    ):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{chr(64 + len(widths))}{max(sheet.max_row, 2)}"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color=WHITE)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = Border(bottom=thin)
                if cell.column >= editable_start:
                    cell.fill = editable_fill
            sheet.row_dimensions[row[0].row].height = 84
        for column_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + column_index)].width = width

    for sheet, validations in (
        (hypothesis_review, [('"1,2,3,4,5"', "N"), ('"Accept,Revise,Reject,Need more evidence"', "O"), ('"Not reviewed,Reviewed,Action required"', "S")]),
        (evidence_review, [('"1,2,3,4,5"', "M"), ('"1,2,3,4,5"', "N"), ('"Yes,Partly,No,Unclear"', "O"), ('"Yes,Partly,No,Unclear"', "P"), ('"Keep,Revise,Remove"', "S"), ('"Not reviewed,Reviewed,Action required"', "T")]),
    ):
        for formula, column in validations:
            validation = DataValidation(
                type="list", formula1=formula, allow_blank=True
            )
            sheet.add_data_validation(validation)
            validation.add(f"{column}2:{column}{max(sheet.max_row, 1000)}")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
